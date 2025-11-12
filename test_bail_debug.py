"""
Debug test pour comprendre pourquoi Article préliminaire ne se génère pas.
"""

import pandas as pd
import openpyxl

print('=' * 80)
print('DEBUG ARTICLE PRÉLIMINAIRE')
print('=' * 80)

# Charger le fichier Excel
wb = openpyxl.load_workbook('Redaction BAIL.xlsx', data_only=True)
ws = wb['Rédaction BAIL']

# Extraire les données pour Article préliminaire
print('\n1️⃣  Données Excel pour Article préliminaire:')
print()

for row_idx in range(1, ws.max_row + 1):
    article = ws.cell(row_idx, 1).value
    if article and 'préliminaire' in str(article).lower():
        print(f'Row {row_idx}: {article}')

        # Colonnes importantes
        designation = ws.cell(row_idx, 2).value
        nom_source = ws.cell(row_idx, 3).value
        source = ws.cell(row_idx, 4).value
        donnee_source = ws.cell(row_idx, 5).value
        condition = ws.cell(row_idx, 6).value
        texte_opt1 = ws.cell(row_idx, 7).value
        condition_opt2 = ws.cell(row_idx, 9).value
        texte_opt2 = ws.cell(row_idx, 10).value

        print(f'  Désignation: {designation}')
        print(f'  Nom Source: {nom_source}')
        print(f'  Source: {source}')
        print(f'  Donnée source: {donnee_source}')
        print(f'  Condition (opt 1): {condition} (type: {type(condition).__name__})')
        print(f'  Texte Option 1: {str(texte_opt1)[:100] if texte_opt1 else "None"}...')
        print(f'  Condition (opt 2): {condition_opt2}')
        print(f'  Texte Option 2: {str(texte_opt2)[:100] if texte_opt2 else "None"}...')
        print()

# Tester avec pandas comme le fait BailGenerator
print('\n2️⃣  Chargement avec pandas (comme BailGenerator):')
df = pd.read_excel('Redaction BAIL.xlsx', sheet_name='Rédaction BAIL')

article_preli = df[df['Article'] == 'Article préliminaire']
print(f'   Lignes trouvées: {len(article_preli)}')
print()

for idx, row in article_preli.iterrows():
    print(f'   Index {idx}:')
    print(f'      Article: {row["Article"]}')
    print(f'      Désignation: {row["Désignation"]}')
    print(f'      Nom Source: {row["Nom Source"]}')
    print(f'      Donnée source: {row["Donnée source"]}')
    print(f'      Condition: {row["Condition"]} (is NaN: {pd.isna(row["Condition"])})')
    print(f'      Condition Option 2: {row["Condition Option 2"]}')
    print()

print('=' * 80)
print('FIN DEBUG')
print('=' * 80)
