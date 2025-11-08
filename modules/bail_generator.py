"""
Générateur de documents BAIL à partir des données et règles Excel.

Ce module gère la logique complexe de génération des baux commerciaux :
- Calcul des variables dérivées
- Évaluation des conditions
- Sélection des variantes d'articles
- Remplacement des placeholders
"""

import pandas as pd
import re
from datetime import datetime, timedelta
from typing import Dict, Optional, List, Any
import logging

logger = logging.getLogger(__name__)


class BailGenerator:
    """Générateur de documents BAIL avec logique conditionnelle."""

    def __init__(self, excel_path: str = "Redaction BAIL.xlsx"):
        """
        Initialise le générateur avec les règles depuis Excel.

        Args:
            excel_path: Chemin vers le fichier Excel contenant les règles
        """
        self.excel_path = excel_path
        self.regles_df = None
        self.donnees_df = None
        self._load_rules()

    def _load_rules(self):
        """Charge les règles depuis le fichier Excel."""
        try:
            import openpyxl

            # Charger avec openpyxl pour lire TOUTES les lignes (même celles avec Article vide)
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            ws_bail = wb["Rédaction BAIL"]

            # Lire toutes les lignes dans une liste de dict
            regles_list = []
            headers = [cell.value for cell in ws_bail[1]]  # Row 1 = headers

            for row_idx in range(2, ws_bail.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, start=1):
                    if header:  # Skip empty headers
                        cell_value = ws_bail.cell(row_idx, col_idx).value
                        row_data[header] = cell_value

                # Ajouter la ligne même si Article est None
                regles_list.append(row_data)

            self.regles_df = pd.DataFrame(regles_list)

            # Charger l'onglet Liste données BAIL
            self.donnees_df = pd.read_excel(
                self.excel_path,
                sheet_name="Liste données BAIL"
            )

            logger.info(f"Règles BAIL chargées: {len(self.regles_df)} lignes")
        except Exception as e:
            logger.error(f"Erreur lors du chargement des règles BAIL: {e}")
            raise

    def _normaliser_nom_variable(self, nom: str) -> str:
        """
        Normalise les noms de variables pour gérer les variations.

        Args:
            nom: Nom de variable brut

        Returns:
            Nom normalisé
        """
        # Mapping des variations de noms
        mappings = {
            "Durée du Bail": "Durée Bail",
            "Durée du DG": "Durée DG",
            "Montant Palier 1": "Montant du palier 1",
            "Montant Palier 2": "Montant du palier 2",
            "Montant Palier 3": "Montant du palier 3",
            "Montant du Palier 1": "Montant du palier 1",
            "Montant du Palier 2": "Montant du palier 2",
            "Montant du Palier 3": "Montant du palier 3",
            "Date prise d'effet": "Date de prise d'effet",
            "Date de prise d'effet du bail": "Date de prise d'effet",
            "Date début bail": "Date de prise d'effet",
            "Date de Prise d'effet + 9 ans": "Date de prise d'effet + 9 ans",
        }

        return mappings.get(nom, nom)

    def calculer_variables_derivees(self, donnees: Dict[str, Any]) -> Dict[str, Any]:
        """
        Calcule les variables dérivées à partir des données primaires.

        Variables calculées :
        - Adresse Locaux Loués
        - Montants des paliers (1 à 6)
        - Surface R-1
        - Type Bail
        - Date de signature
        - Date de prise d'effet + 9 ans
        - Montant du DG
        - Période DG

        Args:
            donnees: Dictionnaire avec les données primaires

        Returns:
            Dictionnaire avec données primaires + dérivées
        """
        # Normaliser les noms de variables en entrée
        derivees = {}
        for key, value in donnees.items():
            normalized_key = self._normaliser_nom_variable(key)
            derivees[normalized_key] = value

        # Adresse Locaux Loués
        ville = derivees.get("Ville ou arrondissement", "")
        rue = derivees.get("Numéro et rue", "")
        if ville and rue:
            derivees["Adresse Locaux Loués"] = f"{ville}, {rue}"

        # Montants des paliers et conversion du loyer
        montant_loyer = None
        montant_loyer_str = derivees.get("Montant du loyer")
        if montant_loyer_str:
            try:
                # Convertir en float (gérer les espaces et virgules)
                montant_loyer_clean = str(montant_loyer_str).replace(" ", "").replace(",", ".")
                montant_loyer = float(montant_loyer_clean)

                for i in range(1, 7):
                    key_annee = f"Loyer année {i}"
                    loyer_annee_str = derivees.get(key_annee)
                    if loyer_annee_str:
                        try:
                            loyer_annee_clean = str(loyer_annee_str).replace(" ", "").replace(",", ".")
                            loyer_annee = float(loyer_annee_clean)
                            derivees[f"Montant du palier {i}"] = montant_loyer - loyer_annee
                        except (ValueError, TypeError):
                            logger.warning(f"Impossible de convertir Loyer année {i}: {loyer_annee_str}")
            except (ValueError, TypeError):
                logger.warning(f"Impossible de convertir Montant du loyer: {montant_loyer_str}")

        # Surface R-1
        surface_totale_str = derivees.get("Surface totale")
        surface_rdc_str = derivees.get("Surface RDC")
        if surface_totale_str and surface_rdc_str:
            try:
                # Convertir en float (gérer les espaces et virgules)
                surface_totale_clean = str(surface_totale_str).replace(" ", "").replace(",", ".")
                surface_rdc_clean = str(surface_rdc_str).replace(" ", "").replace(",", ".")
                surface_totale = float(surface_totale_clean)
                surface_rdc = float(surface_rdc_clean)
                derivees["Surface R-1"] = surface_totale - surface_rdc
            except (ValueError, TypeError):
                logger.warning(f"Impossible de convertir les surfaces: totale={surface_totale_str}, RDC={surface_rdc_str}")

        # Type Bail
        duree_bail_str = derivees.get("Durée Bail")
        if duree_bail_str:
            try:
                duree_bail = int(float(str(duree_bail_str).replace(" ", "").replace(",", ".")))
                if duree_bail == 9:
                    derivees["Type Bail"] = "3/6/9"
                elif duree_bail == 10:
                    derivees["Type Bail"] = "6/9/10"
            except (ValueError, TypeError):
                logger.warning(f"Impossible de convertir Durée Bail: {duree_bail_str}")

        # Date de signature (aujourd'hui + 15 jours)
        date_signature = datetime.now() + timedelta(days=15)
        derivees["Date de signature"] = date_signature.strftime("%d/%m/%Y")

        # Date de prise d'effet + 9 ans
        date_prise_effet_str = derivees.get("Date de prise d'effet")
        if date_prise_effet_str:
            try:
                # Parser la date (format DD/MM/YYYY ou DD-MM-YYYY ou autres formats courants)
                for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y"]:
                    try:
                        date_prise_effet = datetime.strptime(str(date_prise_effet_str).strip(), fmt)
                        # Ajouter 9 ans
                        date_fin_bail = date_prise_effet + timedelta(days=365*9)
                        date_str = date_fin_bail.strftime("%d/%m/%Y")
                        # Ajouter les deux variantes de casse
                        derivees["Date de prise d'effet + 9 ans"] = date_str
                        derivees["Date de Prise d'effet + 9 ans"] = date_str
                        logger.debug(f"Date de prise d'effet + 9 ans calculée: {date_str}")
                        break
                    except ValueError:
                        continue
            except Exception as e:
                logger.warning(f"Impossible de calculer Date de prise d'effet + 9 ans: {e}")

        # Montant du DG
        duree_dg_str = derivees.get("Durée DG")
        duree_dg = None
        if duree_dg_str:
            try:
                duree_dg = float(str(duree_dg_str).replace(" ", "").replace(",", "."))
            except (ValueError, TypeError):
                logger.warning(f"Impossible de convertir Durée DG: {duree_dg_str}")

        if montant_loyer and duree_dg:
            derivees["Montant du DG"] = (montant_loyer / 12) * duree_dg

        # Période DG
        if duree_dg:
            periode_map = {3: "quart", 4: "tiers", 6: "moitier"}
            derivees["Période DG"] = periode_map.get(int(duree_dg), "")

        logger.debug(f"Variables dérivées calculées: {list(derivees.keys())}")
        return derivees

    def evaluer_condition(self, condition_str: str, donnees: Dict[str, Any]) -> bool:
        """
        Évalue une condition textuelle.

        Exemples de conditions:
        - "Si [Durée Bail] > 9"
        - "Si [Actualisation] = 'Oui'"
        - "Si [Loyer année 1] non vide"
        - "Si plusieurs conditions suspensives"
        - None ou vide → True (pas de condition)

        Args:
            condition_str: Condition en format texte
            donnees: Données pour évaluer la condition

        Returns:
            True si la condition est satisfaite, False sinon
        """
        # Pas de condition = toujours vrai
        if pd.isna(condition_str) or not condition_str:
            return True

        condition = str(condition_str).strip()

        # Cas spécial: "Si plusieurs conditions suspensives"
        if "plusieurs conditions suspensives" in condition.lower():
            count = sum(1 for i in range(1, 5)
                       if donnees.get(f"Condition suspensive {i}"))
            return count > 1

        # Nettoyer les guillemets typographiques
        condition = condition.replace('"', '"').replace('"', '"').replace(''', "'").replace(''', "'")

        # Parser les conditions avec pattern "Si [Variable] opérateur valeur"
        # Pattern: Si [Variable] (=|>|<|>=|<=|!=|supérieur à) valeur
        match_comparison = re.search(
            r'Si\s+"?([^"\[\]]+|[\[][^\]]+[\]])"?\s*(=|>|<|>=|<=|!=|supérieur à|supérieure à)\s*["\']?([^"\']+)["\']?',
            condition,
            re.IGNORECASE
        )

        if match_comparison:
            var_name = match_comparison.group(1).strip().replace('[', '').replace(']', '')
            var_name = self._normaliser_nom_variable(var_name)
            operator = match_comparison.group(2).strip().lower()
            expected_value = match_comparison.group(3).strip()

            actual_value = donnees.get(var_name)

            # Gérer les comparaisons
            try:
                if operator == "=":
                    return str(actual_value).strip() == expected_value
                elif operator == "!=":
                    return str(actual_value).strip() != expected_value
                elif operator in [">", "supérieur à", "supérieure à"]:
                    return float(actual_value) > float(expected_value)
                elif operator == ">=":
                    return float(actual_value) >= float(expected_value)
                elif operator == "<":
                    return float(actual_value) < float(expected_value)
                elif operator == "<=":
                    return float(actual_value) <= float(expected_value)
            except (ValueError, TypeError):
                logger.warning(f"Impossible de comparer {actual_value} avec {expected_value}")
                return False

        # Pattern: Si [Variable] non vide / non nul
        match_nonempty = re.search(
            r'Si\s+\[([^\]]+)\]\s+non\s+(vide|nul)',
            condition,
            re.IGNORECASE
        )

        if match_nonempty:
            var_name = match_nonempty.group(1).strip()
            value = donnees.get(var_name)
            # Considérer comme non vide si value existe et n'est pas None, "", 0, False
            return bool(value) and value != 0

        # Si on ne peut pas parser, logger un warning
        logger.warning(f"Condition non reconnue: {condition}")
        return False

    def obtenir_texte_article(
        self,
        article_name: str,
        designation: Optional[str],
        donnees: Dict[str, Any]
    ) -> Optional[str]:
        """
        Obtient le texte d'un article en évaluant les conditions.

        Args:
            article_name: Nom de l'article (ex: "Comparution")
            designation: Désignation spécifique (ex: "Comparution Bailleur")
            donnees: Données pour évaluer les conditions

        Returns:
            Texte de l'article ou None si non trouvé
        """
        # NOUVELLE LOGIQUE: Chercher d'abord la ligne avec Article + Désignation
        # puis inclure toutes les lignes suivantes jusqu'au prochain Article non-null

        lignes_candidates = []
        found_start = False
        current_designation = None

        for idx, row in self.regles_df.iterrows():
            article_val = row['Article']
            designation_val = row['Désignation']

            # Nouvelle section d'article
            if pd.notna(article_val):
                # Si on était déjà dans notre section, on s'arrête
                if found_start and article_val != article_name:
                    break

                # Vérifier si c'est le début de notre article
                if article_val == article_name:
                    if designation is None or designation_val == designation:
                        found_start = True
                        current_designation = designation_val
                        lignes_candidates.append(row)
                    elif found_start:
                        # Nouvelle désignation du même article, on s'arrête
                        break
                else:
                    found_start = False

            # Ligne de continuation (Article = None)
            elif found_start:
                lignes_candidates.append(row)

        if not lignes_candidates:
            logger.warning(f"Aucune règle trouvée pour l'article '{article_name}'")
            return None

        # Parcourir les lignes candidates et évaluer les conditions
        for ligne in lignes_candidates:
            # Vérifier si la donnée source correspond (pour les lookup tables)
            donnee_source = ligne.get('Donnée source')
            nom_source = ligne.get('Nom Source')

            if pd.notna(donnee_source) and pd.notna(nom_source):
                # C'est un lookup: vérifier si la valeur correspond
                # Gérer le cas où Nom Source contient plusieurs variables (multiligne)
                noms_sources = [n.strip() for n in str(nom_source).split('\n') if n.strip()]

                match_found = False
                for nom in noms_sources:
                    valeur_actuelle = donnees.get(nom)
                    if str(valeur_actuelle) == str(donnee_source):
                        match_found = True
                        break

                if not match_found:
                    continue  # Passer à la ligne suivante

            # Évaluer Condition Option 1
            condition1 = ligne.get('Condition')
            if self.evaluer_condition(condition1, donnees):
                texte = ligne.get('Entrée correspondante - Option 1')
                if pd.notna(texte):
                    return str(texte)

            # Évaluer Condition Option 2
            condition2 = ligne.get('Condition Option 2')
            if self.evaluer_condition(condition2, donnees):
                texte = ligne.get('Entrée correspondante - Option 2')
                if pd.notna(texte):
                    return str(texte)

        logger.warning(f"Aucune condition satisfaite pour l'article '{article_name}'")
        return None

    def remplacer_placeholders(self, texte: str, donnees: Dict[str, Any]) -> str:
        """
        Remplace les placeholders [Variable] dans le texte.

        Args:
            texte: Texte avec placeholders
            donnees: Données pour remplacer les placeholders

        Returns:
            Texte avec placeholders remplacés
        """
        if not texte:
            return ""

        # Trouver tous les placeholders [Variable]
        placeholders = re.findall(r'\[([^\]]+)\]', texte)

        for placeholder in placeholders:
            # Normaliser le nom de la variable
            placeholder_norm = self._normaliser_nom_variable(placeholder)
            valeur = donnees.get(placeholder_norm) or donnees.get(placeholder)
            if valeur is not None:
                # Formater les nombres avec séparateurs si nécessaire
                if isinstance(valeur, (int, float)):
                    valeur_str = f"{valeur:,.2f}".replace(",", " ").replace(".", ",")
                    # Retirer les décimales si .00
                    valeur_str = valeur_str.replace(",00", "")
                else:
                    valeur_str = str(valeur)

                texte = texte.replace(f"[{placeholder}]", valeur_str)
            else:
                logger.warning(f"Placeholder non trouvé dans les données: [{placeholder}]")
                texte = texte.replace(f"[{placeholder}]", f"[{placeholder}]")  # Garder le placeholder

        return texte

    def generer_bail(self, donnees: Dict[str, Any]) -> Dict[str, str]:
        """
        Génère le contenu complet du BAIL.

        Args:
            donnees: Dictionnaire avec toutes les données primaires

        Returns:
            Dictionnaire avec les articles générés
            {
                "Comparution": "texte...",
                "Article préliminaire": "texte...",
                ...
            }
        """
        logger.info("Début de la génération du BAIL")

        # 1. Calculer les variables dérivées
        donnees_complete = self.calculer_variables_derivees(donnees)

        # 2. Générer chaque article
        articles_generes = {}

        # Liste des articles à générer (dans l'ordre)
        articles_order = [
            ("Comparution", "Comparution Bailleur"),
            ("Comparution", "Comparution Preneur"),
            ("Article préliminaire", None),
            ("Article 1", None),
            ("Article 2", None),
            ("Article 3", None),
            ("Article 5.3", None),
            ("Article 7.1", None),
            ("Article 7.2", None),
            ("Article 7.3", None),
            ("Article 7.6", None),
            ("Article 8", None),
            ("Article 19", None),
            ("Article 22.2", None),
            ("Article 26", None),
            ("Article 26.1", None),
            ("Article 26.2", None)
        ]

        for item in articles_order:
            # item peut être un tuple (article_name, designation) ou juste article_name
            if isinstance(item, tuple):
                article_name, designation = item
            else:
                article_name = item
                designation = None

            texte = self.obtenir_texte_article(article_name, designation, donnees_complete)

            if texte:
                # Remplacer les placeholders
                texte_final = self.remplacer_placeholders(texte, donnees_complete)

                # Clé: utiliser designation si présente, sinon article_name
                key = designation if designation else article_name
                articles_generes[key] = texte_final
                logger.debug(f"Article généré: {key}")
            else:
                logger.warning(f"Article non généré: {article_name} (designation={designation})")

        logger.info(f"Génération terminée: {len(articles_generes)} articles générés")
        return articles_generes
