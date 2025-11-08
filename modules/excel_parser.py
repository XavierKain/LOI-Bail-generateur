"""
Module pour extraire les données des fichiers Excel (Fiche de décision).
Lit les données depuis les différents onglets et les map vers les variables LOI.
"""

import logging
from typing import Dict, Optional
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)


class ExcelParser:
    """Parse les fichiers Excel de décision pour extraire les variables LOI."""

    def __init__(self, excel_path: str, config_path: str = "Rédaction LOI.xlsx"):
        """
        Initialise le parser avec le fichier Excel source.

        Args:
            excel_path: Chemin vers le fichier Excel source (Fiche de décision)
            config_path: Chemin vers le fichier de configuration (Rédaction LOI.xlsx)
        """
        self.excel_path = Path(excel_path)
        self.config_path = Path(config_path)

        if not self.excel_path.exists():
            raise FileNotFoundError(f"Fichier Excel source introuvable: {excel_path}")
        if not self.config_path.exists():
            raise FileNotFoundError(f"Fichier de configuration introuvable: {config_path}")

        try:
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            self.config_workbook = openpyxl.load_workbook(self.config_path, data_only=True)
            # Also load config with formulas to handle cases where cached values are missing
            self.config_workbook_formulas = openpyxl.load_workbook(self.config_path, data_only=False)
            logger.info(f"Fichier Excel chargé: {self.excel_path.name}")
            logger.info(f"Configuration chargée: {self.config_path.name}")
        except InvalidFileException as e:
            raise ValueError(f"Fichier Excel invalide: {e}")

    def _get_cell_value(self, sheet_name: str, cell_ref: str) -> Optional[str]:
        """
        Récupère la valeur d'une cellule depuis un onglet.

        Args:
            sheet_name: Nom de l'onglet
            cell_ref: Référence de la cellule (ex: "B23")

        Returns:
            Valeur de la cellule ou None
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                logger.warning(f"Onglet '{sheet_name}' introuvable")
                return None

            sheet = self.workbook[sheet_name]
            value = sheet[cell_ref].value

            # Convertir les valeurs en string, gérer les dates
            if value is None:
                return None
            elif isinstance(value, datetime):
                return value.strftime("%d/%m/%Y")
            elif isinstance(value, (int, float)):
                return str(value)
            else:
                return str(value).strip()

        except Exception as e:
            logger.warning(f"Erreur lecture cellule {sheet_name}!{cell_ref}: {e}")
            return None

    def _parse_formula(self, formula: str) -> Optional[str]:
        """
        Parse une formule Excel pour extraire la valeur.

        Args:
            formula: Formule Excel (ex: "=Validation!B23")

        Returns:
            Valeur extraite ou None
        """
        if not formula or not isinstance(formula, str):
            return None

        # Retirer le signe =
        formula = formula.strip()
        if formula.startswith("="):
            formula = formula[1:]

        # Format: 'Sheet Name'!CellRef ou SheetName!CellRef
        if "!" in formula:
            parts = formula.split("!")
            sheet_name = parts[0].strip("'")
            cell_ref = parts[1].strip()
            return self._get_cell_value(sheet_name, cell_ref)

        return None

    def extract_variables(self) -> Dict[str, str]:
        """
        Extrait toutes les variables depuis le fichier Excel source.
        Utilise le fichier de configuration pour savoir quoi extraire.

        Returns:
            Dictionnaire {nom_variable: valeur}
        """
        variables = {}

        # Lire la configuration depuis Rédaction LOI
        config_sheet = self.config_workbook["Rédaction LOI"]
        config_sheet_formulas = self.config_workbook_formulas["Rédaction LOI"]

        # Parcourir les lignes de configuration (ligne 2 à 40+)
        for row in range(2, max(config_sheet.max_row, config_sheet_formulas.max_row) + 1):
            nom = config_sheet.cell(row, 1).value  # Colonne A: Nom
            source = config_sheet.cell(row, 2).value  # Colonne B: Source

            # If source is None, try getting the formula
            if not source:
                source = config_sheet_formulas.cell(row, 2).value

            if not nom:
                continue

            nom = str(nom).strip()

            # Cas spéciaux: formules de calcul dans la config
            if source and isinstance(source, str):
                if source.startswith("=") and "!" in source:
                    # C'est une référence à une cellule
                    value = self._parse_formula(source)
                    if value:
                        variables[nom] = value
                elif "[" in source and "]" in source:
                    # C'est une formule qui sera calculée plus tard (ex: adresse, paliers)
                    # On la stocke pour traitement ultérieur
                    variables[f"_formula_{nom}"] = source
                else:
                    # Texte littéral ou description
                    variables[f"_description_{nom}"] = source

        # Ajouter la date d'aujourd'hui
        variables["Date d'aujourd'hui"] = datetime.now().strftime("%d/%m/%Y")

        logger.info(f"{len(variables)} variables extraites")
        return variables

    def extract_societe_info(self) -> Dict[str, Dict[str, str]]:
        """
        Extrait les informations des sociétés bailleures depuis la configuration.

        Returns:
            Dictionnaire {nom_societe: {header: str, footer: str}}
        """
        societes = {}

        config_sheet = self.config_workbook["Société Bailleur"]

        # Parcourir les lignes (ligne 2 = première société)
        for row in range(2, config_sheet.max_row + 1):
            nom_societe = config_sheet.cell(row, 1).value  # Colonne A
            header = config_sheet.cell(row, 2).value  # Colonne B
            footer = config_sheet.cell(row, 3).value  # Colonne C

            if not nom_societe:
                continue

            nom_societe = str(nom_societe).strip()

            societes[nom_societe] = {
                "header": str(header).strip() if header else nom_societe,
                "footer": str(footer).strip() if footer else ""
            }

        logger.info(f"{len(societes)} sociétés bailleures chargées")
        return societes

    def get_output_filename(self, variables: Dict[str, str]) -> str:
        """
        Génère le nom du fichier de sortie basé sur les variables extraites.
        Format: "YYYY MM DD - LOI NomPreneur.docx"

        Args:
            variables: Dictionnaire des variables

        Returns:
            Nom du fichier de sortie
        """
        date_loi = variables.get("Date LOI", "")
        nom_preneur = variables.get("Nom Preneur", "INCONNU")

        # Parser la date si elle existe
        if date_loi:
            try:
                # Format attendu: DD/MM/YYYY
                if "/" in date_loi:
                    parts = date_loi.split("/")
                    date_str = f"{parts[2]} {parts[1]} {parts[0]}"
                else:
                    # Utiliser la date d'aujourd'hui
                    date_str = datetime.now().strftime("%Y %m %d")
            except:
                date_str = datetime.now().strftime("%Y %m %d")
        else:
            date_str = datetime.now().strftime("%Y %m %d")

        filename = f"{date_str} - LOI {nom_preneur}.docx"
        return filename
