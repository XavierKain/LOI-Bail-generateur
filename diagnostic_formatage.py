"""
Script de diagnostic pour comprendre pourquoi le formatage ne fonctionne pas.
"""

import openpyxl
import sys

print("=" * 80)
print("DIAGNOSTIC COMPLET DU FORMATAGE")
print("=" * 80)
print()

# ÉTAPE 1: Vérifier le fichier Excel
print("1. VÉRIFICATION DU FICHIER EXCEL 'Redaction BAIL.xlsx'")
print("-" * 80)

try:
    wb = openpyxl.load_workbook('Redaction BAIL.xlsx', data_only=True)
    ws = wb['Rédaction BAIL']

    print("✅ Fichier Excel chargé avec succès")
    print()

    # Vérifier les lignes avec balises
    lignes_avec_balises = []
    for row_idx in range(2, 20):
        nom_source = ws.cell(row=row_idx, column=3).value
        donnee_source = ws.cell(row=row_idx, column=5).value
        option1 = ws.cell(row=row_idx, column=7).value

        if option1 and '<b>' in str(option1):
            lignes_avec_balises.append({
                'ligne': row_idx,
                'nom_source': nom_source,
                'donnee_source': donnee_source,
                'texte': str(option1)[:100]
            })

    if lignes_avec_balises:
        print(f"✅ Trouvé {len(lignes_avec_balises)} ligne(s) avec balises <b>:")
        for info in lignes_avec_balises:
            print(f"   Ligne {info['ligne']}: Donnée source = '{info['donnee_source']}'")
            print(f"   Texte: {info['texte']}...")
            print()
    else:
        print("❌ AUCUNE ligne avec balises <b> trouvée dans 'Redaction BAIL.xlsx'!")
        print("   Problème: Le fichier Excel n'a pas de balises de formatage")
        sys.exit(1)

except Exception as e:
    print(f"❌ Erreur lors de la lecture du fichier Excel: {e}")
    sys.exit(1)

# ÉTAPE 2: Simuler la génération avec une donnée de test
print()
print("2. SIMULATION DE LA GÉNÉRATION")
print("-" * 80)

from modules.bail_generator import BailGenerator
from modules.bail_word_generator import BailWordGenerator

# Test avec SCI HSR 2 (ligne 5 qui a les balises)
donnees_test = {
    'Société Bailleur': 'SCI HSR 2'
}

print(f"Test avec: Société Bailleur = 'SCI HSR 2'")
print()

try:
    generator = BailGenerator("Redaction BAIL.xlsx", None)

    # Obtenir le texte pour Comparution Bailleur
    texte = generator.obtenir_texte_article("Comparution", "Comparution Bailleur", donnees_test)

    if texte:
        print(f"✅ Texte généré pour Comparution Bailleur:")
        print(f"   Longueur: {len(texte)} caractères")
        print(f"   Contient <b>: {'<b>' in texte}")
        print(f"   Premiers 200 caractères:")
        print(f"   {texte[:200]}")
        print()

        if '<b>' in texte:
            print("✅ Les balises <b> SONT présentes dans le texte généré")
            print()

            # ÉTAPE 3: Tester le parser
            print("3. TEST DU PARSER DE BALISES")
            print("-" * 80)

            segments = BailWordGenerator._parse_formatting_tags(texte[:300])
            print(f"Parsing des 300 premiers caractères:")
            for i, (text, formatting) in enumerate(segments):
                if i < 5:  # Montrer les 5 premiers segments
                    print(f"  Segment {i}:")
                    print(f"    Texte: {text[:60]}...")
                    print(f"    Format: {formatting}")
            print()

            print("✅ Le parser fonctionne correctement")

        else:
            print("❌ PROBLÈME: Les balises <b> NE SONT PAS dans le texte généré!")
            print("   Raison possible: bail_generator.py ne retourne pas la bonne colonne")

    else:
        print("❌ PROBLÈME: Aucun texte généré pour Comparution Bailleur")
        print("   Raison possible: La condition ne matche pas ou la donnée source ne correspond pas")

except Exception as e:
    print(f"❌ Erreur lors de la génération: {e}")
    import traceback
    traceback.print_exc()

print()
print("=" * 80)
print("FIN DU DIAGNOSTIC")
print("=" * 80)
